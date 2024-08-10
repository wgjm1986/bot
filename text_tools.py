import docx, pptx, odf, csv, re, json, spacy
import boto3
from PIL import Image
from io import BytesIO
from openpyxl import load_workbook
from odf.opendocument import load
from odf import text
from odf import teletype

from pypdf import PdfReader
from pathlib import Path

# Retrieve the text of a given document based on file path.
def get_document_paragraphs(file_path,last_pdf_page=None):
    extension = Path(file_path).suffix
    if extension == ".txt" or extension == ".tex":
        with open(file_path, 'r') as file:
            document_text = file.read()
            document_paragraphs = document_text.split('\n\n')
            document_paragraphs = [para for para in document_paragraphs if para]
            return document_paragraphs
    elif extension == ".ipynb":
        with open(file_path, 'r') as file:
            notebook_json = json.load(file)
        processed_cells = []
        for cell in notebook_json['cells']:
            # continue if cell_type is not cells or markdown?
            cell_content = ''.join(cell['source'])
            processed_cells.append(cell_content)
            if cell['cell_type'] == 'code' and 'outputs' in cell:
                for output in cell['outputs']:
                    if 'data' in output:
                        for mime_type, data in output['data'].items():
                            if mime_type.startswith('image/png') or mime_type.startswith('image/jpeg'):
                                processed_cells.append('[Image content removed]')
                            elif mime_type.startswith('application/pdf'):
                                processed_cells.append('[PDF content removed]')
                            else:
                                processed_cells.append(''.join(data))
        document_paragraphs = processed_cells
        return document_paragraphs
    elif extension == ".pdf":
        document_text = ""
        pdf = PdfReader(file_path)
        pages = pdf.pages[0:last_pdf_page]
        pages_text = [ page.extract_text() for page in pages ]
        document_paragraphs = [para for page in pages_text for para in page.split('\n\n')]
        return document_paragraphs
    elif extension == ".docx":
        document_paragraphs = docx.Document(file_path).paragraphs
        return document_paragraphs
    elif extension == ".pptx":
        textract = boto3.client('textract',region_name='us-east-1')
        presentation = pptx.Presentation(file_path)
        document_paragraphs = []
        for slide in presentation.slides:
            slide_shapes_text = []
            for shape in slide.shapes:
                if hasattr(shape,"text"):
                    slide_shapes_text.append(shape.text)
                # if shape.shape_type == pptx.enum.shapes.MSO_SHAPE_TYPE.PICTURE:
                #     image_stream = shape.image.blob
                #     image = Image.open(BytesIO(image_stream))
                #     buffered = BytesIO()
                #     image.save(buffered,format="PNG")
                #     image_bytes = buffered.getvalue()
                #     response = textract.detect_document_text(Document={'Bytes':image_bytes})
                #     image_text = []
                #     for item in response['Blocks']:
                #         if item['BlockType'] == 'LINE':
                #             image_text.append(item['Text'])
                #     if image_text.strip():
                #         slide_shapes_text.append(image_text)
            slide_text = '\n'.join(slide_shapes_text)
            document_paragraphs.append(slide_text)
        return document_paragraphs
    elif extension == ".xlsx":
        workbook_values = load_workbook(filename=file_path,data_only=True)
        # with data_only=False, cell.value will be a formula when there is one (otherwise a value as above)
        workbook_formulas = load_workbook(filename=file_path,data_only=False)
        sheets = []
        for sheet_name in workbook_values.sheetnames:
            output_lines = []
            sheet_values = workbook_values[sheet_name]
            sheet_formulas = workbook_formulas[sheet_name]
            output_lines.append(f"This is a sheet from an xlsx file. The sheet name is {sheet_name}.\nEach non-empty cell in the sheet is listed below along with its value, and its formula if any.")
            for row in sheet_values.iter_rows():
                row_data = []
                for cell in row:
                    cell_value = cell.value
                    cell_formula = sheet_formulas[cell.coordinate].value
                    if cell_value is None and cell_formula is None: continue
                    cell_label = f"{cell.coordinate}: Value: {cell_value}"
                    if cell_formula != cell_value: cell_label += f", Formula: {cell_formula}"
                    row_data.append(cell_label)
                if row_data:
                    output_lines.append(" | ".join(str(item) if item is not None else '' for item in row_data))
            if output_lines: sheets.append('\n'.join(output_lines))
        # For now, treat each sheet as a paragraph. May need to modify this depending on how big the output is
        return sheets
    elif extension == ".odt":
        document = load(file_path)
        document_paragraphs_raw = document.getElementsByType(text.P)
        document_paragraphs = [ teletype.extractText(para) for para in document_paragraphs_raw ]
        return document_paragraphs
    else:
        print("Unsupported file type: " + file_path)
        return

# Chunk raw text from a document
def chunk_paragraphs(paragraphs):
    chunk_size = 3
    overlap = 0
    chunks = []
    i = 0
    while i < len(paragraphs):
        end_index = i + chunk_size
        chunk = paragraphs[i:end_index]
        for j,paragraph in enumerate(chunk):
            # Replace tabs with four spaces
            chunk[j] = re.sub('\t','    ',paragraph)
            # Allow only four spaces max at any time
            chunk[j] = re.sub('    +','    ',paragraph)
            # Allow only one newline at a time
            chunk[j] = re.sub('\n+','\n',paragraph)
        chunk_text = '\n\n'.join(chunk)
        if re.search('[A-Za-z]',chunk_text): chunks.append(chunk_text)
        i += chunk_size - overlap
    return chunks
